# -*- coding: utf-8 -*-
"""
Created on Mon Dec 14 09:16:28 2020

@author: serw12
"""

import csv
import matplotlib.pyplot as plt
from matplotlib.legend_handler import HandlerLine2D

import openpyxl
from pathlib import Path
from datetime import datetime
import numpy as np #Helps me transposing lists of lists similar to a matrix.
import math

filename1 = 'Costomarginal_S1_Dic2020.xlsx'
filename2 = 'Costomarginal_S2_Dic2020.xlsx'
filename3 = 'Costomarginal_S3_Dic2020.xlsx'
filename4 = 'Costomarginal_S4_Dic2020.xlsx'
filename5 = 'Costomarginal_S5_Dic2020.xlsx'
Filenames = [filename1,filename2,filename3,filename4,filename5]
Plant = 22
# Plant hace referencia a la planta generadora 22 == Cañaveral.

#for dat in range(1,6):
xlsx_file = Path("Data",filename1)
wb_obj = openpyxl.load_workbook(xlsx_file) 
# Read the active sheet:
sheet = wb_obj.active
maxrow = sheet.max_row
maxcolumn = sheet.max_column
#Idnetify size of the worksheet.      
print ("Estación de interes:",sheet.cell(1,Plant).value)
# cell(#,#) identifica la celda en excel e.j cell(A,22).
timevec = []
names = []
pricevec = []
pricestats = []
pricemean = []
testvec = []
i = 3 #Prices start from the 3th row.
j = 2 #Prices start from the second column.
lm=0
while i <= maxrow:
    time = sheet.cell(i,1).value
    timevec.append(time)
    while j <= maxcolumn:
        price = sheet.cell(i,j).value
        pricevec.append(price)
        j=j+1
        lm=lm+1    
    pricestats.append(pricevec)
    pricevec = []
    j=2
    i = i+1 
            #Prices as stored as: pricestats[row][column] 
            #Where row==Plant and column==time           
partsigma = []
pricesigma = []             
for m in pricestats:
   mean = sum(m)/maxcolumn
   for x in m:
        partsigma.append((x-mean)**2)        
        sigma = math.sqrt(sum(partsigma)/maxcolumn)
   partsigma = []
   pricemean.append(mean)
   pricesigma.append(sigma)                       
pricehist = np.array(pricestats).T.tolist()
    
print(timevec[1])
plt.plot(pricemean)
plt.plot(pricehist[Plant])
plt.plot(pricesigma)
plt.show()


plt.figure(figsize=(10,5))
Media, = plt.plot(pricemean,Label="Media Golbal")
priceplant, = plt.plot(pricehist[Plant],Label=sheet.cell(1,Plant).value)
sigma, = plt.plot(pricesigma,Label="Desviación est.")
plt.legend(handler_map={Media: HandlerLine2D(numpoints=4)})
plt.show()





#with open(filename) as f:
#    reader = csv.reader(f)
#    header_now = next(reader)
#    print(header_now)
    
    
    
    
