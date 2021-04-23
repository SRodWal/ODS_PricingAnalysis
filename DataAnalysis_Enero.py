# -*- coding: utf-8 -*-
"""
Created on Mon Dec 14 09:16:28 2020

@author: serw12
"""

import csv
import matplotlib.pyplot as plt
from matplotlib.legend_handler import HandlerLine2D

import openpyxl
from pathlib import Path
from datetime import datetime
import numpy as np #Helps me transposing lists of lists similar to a matrix.
import math

filename1 = 'CM_Ene_S1.xlsx'
filename2 = 'CM_Ene_S2.xlsx'
filename3 = 'CM_Ene_S3.xlsx'
filename4 = 'CM_Ene_S4.xlsx'
filename5 = "CM_Feb_S1.xlsx"
Filenames = [filename1,filename2,filename3,filename4,filename5]
Canaveral = 22
Guaimas = 52
Plant = Guaimas
# Plant hace referencia a la planta generadora 22 == Cañaveral.
pricehiststat=[]
globalmean=[]
storemean=[]
plantmean=[]
for dat in range(0,len(Filenames)):
    print("Semana ",dat)
    xlsx_file = Path("Data",Filenames[dat])
    wb_obj = openpyxl.load_workbook(xlsx_file)
    # Read the active sheet:
    sheet = wb_obj.active
    maxrow = sheet.max_row
    maxcolumn = sheet.max_column
    #Idnetify size of the worksheet.  
    print("Numero de plantas generadoras:",maxcolumn-1)    
    print ("Estación de interes:",sheet.cell(1,Plant).value)
    # cell(#,#) identifica la celda en excel e.j cell(A,22).
    timevec = []
    names = []
    pricevec = []
    pricestats = []
    pricemean = []
    testvec = []
    i = 3 #Prices start from the 3th row.
    j = 2 #Prices start from the second column.
    lm=0
    while i <= maxrow:
        time = sheet.cell(i,1).value
        timevec.append(time)
        while j <= maxcolumn:
            price = sheet.cell(i,j).value
            pricevec.append(price)
            j=j+1
            lm=lm+1    
        pricestats.append(pricevec)
        pricevec = []
        j=2
        i = i+1 
            #Prices as stored as: pricestats[row][column] 
            #Where row==time and column==Plant          
    partsigma = []
    pricesigma = []             
    for m in pricestats:
        mean = sum(m)/maxcolumn
        for x in m:
            partsigma.append((x-mean)**2)        
            sigma = math.sqrt(sum(partsigma)/maxcolumn)
        partsigma = []
        pricemean.append(mean)
        pricesigma.append(sigma)                       
    pricehist = np.array(pricestats).T.tolist()
    
    print(timevec[0])
    plt.figure(figsize=(10,5))
    Media, = plt.plot(pricemean,Label="Media Golbal")
    priceplant, = plt.plot(pricehist[Plant],Label=sheet.cell(1,Plant).value)
    sigma, = plt.plot(pricesigma,Label="Desviación est.")
    plt.legend(handler_map={Media: HandlerLine2D(numpoints=4)})
    plt.title("Precios: semana del "+str(timevec[0].date())+" al "+str(timevec[167].date()))
    plt.show()
    pricedif=[]
    for m in range(0,len(pricemean)):
        dif = 200*(pricehist[Plant][m]-pricemean[m])/(pricehist[Plant][m]+pricemean[m])
        pricedif.append(dif)
    plt.figure(figsize=(10,5))    
    plt.plot(pricedif)
    plt.title("Diferencia %: "+str(timevec[0].date())+" al "+str(timevec[167].date()))
    plt.show()
    pricehiststat.extend(pricehist)
    storemean.append(pricemean)
    plantmean.append(pricehist[Plant])
    
for t in range(0,168):
    mean = sum(pricehiststat[0:len(pricehiststat)-1][t])/len(pricehiststat)
    globalmean.append(mean)  

globalM=[]
globaldiv=[]
plantM=[]
plantdiv=[]
newstoremean = np.array(storemean).T.tolist() #Re-arrange dataset for analysis
newplantmean = np.array(plantmean).T.tolist() #Re-arrange dataset for analysis
for t in newstoremean:
    mean = sum(t)/len(t)
    globalM.append(mean)
    globaldiv.append(np.std(t))    
for t in newplantmean:
    mean = sum(t)/len(t)
    plantM.append(mean)
    plantdiv.append(np.std(t))
plt.figure(figsize=(10,5))   
globM,= plt.plot(globalM,Label="Media global")
globdiv, = plt.plot(globaldiv,Label="Desviacion est. global")
plntM,= plt.plot(plantM,Label="Media de "+str(sheet.cell(1,Plant).value))
plntdiv,= plt.plot(plantdiv,Label="Desvicacion est. "+str(sheet.cell(1,Plant).value))
plt.legend(handler_map={globM: HandlerLine2D(numpoints=4)})
plt.title("Promedio global semanal")
plt.show

plt.figure(figsize=(10,5))
lu,=plt.plot(globalM[0:23],Label="Lunes")
ma,=plt.plot(globalM[24:47],Label="Martes")
mi,=plt.plot(globalM[48:71],Label="Miercoles")
ju,=plt.plot(globalM[72:95],Label="Jueves")
vi,=plt.plot(globalM[96:119],Label="Viernes")
sa,=plt.plot(globalM[120:143],Label="Sabado")
do,=plt.plot(globalM[144:172],Label="Domingo")
plt.legend(handler_map={lu: HandlerLine2D(numpoints=7)})
plt.show


    
    


#with open(filename) as f:
#    reader = csv.reader(f)
#    header_now = next(reader)
#    print(header_now)
    
    
    
    
